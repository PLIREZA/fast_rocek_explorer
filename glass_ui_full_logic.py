import sys
import threading
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout,
    QHBoxLayout, QFileDialog, QComboBox, QTabWidget, QFrame, QSizePolicy, QProgressBar
)
from PyQt5.QtGui import QPainter, QBrush, QColor, QCursor
from PyQt5.QtCore import Qt, QPoint, pyqtSignal, QObject
from warnings import filterwarnings

filterwarnings('ignore')

# Lazy imports for heavy libraries
def import_processing_libs():
    import chardet
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Alignment
    return chardet, pd, load_workbook, PatternFill, Alignment

# FCR settings
remove_ids = {
    "78bdc1ebc25161b36158121a44ed4d3c", "8aca6bae7e2801fddb4a5a51ebc876a8", "6df46ae9448c69382aa21223d11b5408",
    "b7616efa3a30df2150659f2a9bb02e94", "d361816954ae755f48273b21f6af721e", "868e5aa123badb05c0abcd3266609844",
    "7a63504eaed7e0142c538565f255c890", "524e8d3a781344e3d8919de55371dc3e", "cc3d38af3f82d2e2939b26444a67b9a5",
    "48e892ae0eef979c5c6f3038ba499bb9", "5c794a6d8be5a107f413bc963894ddc6", "88cfdd78026fdb7757c10bf3b17612d3",
    "807ad92c313df42e2e62401938ba8ff1", "c5cafed8a9d21286a704d4f450e1a24e", "aa5b64e98cb6dae12f6aff71c5713987",
    "eca7b79e62502fc5e5aff69335ffcc36", "493eeb9c634fb669812534e25ea94fcf", "2fb2599afae6c6e2c7bcc3b47089026a",
    "e9e5d41ad172a6b25559c86ef459c33b", "6c944d4960a1af53aa68b6dbaa9f7de5", "c5d79a6a2336be984d2f5e407d6bec79",
    "23f687554ce87909fcfd3c7b5a1f8611", "67bb555f6b59b807880178cc5e29938f", "4d692967b23339b592ed802e6741a7f5",
    "ac54967ce15abb03ce7852381b1c0da2", "5351d7ac52fc42d90ffb7c4b4a338177", "bc242e96f5505f79ded07a97b7c820aa",
    "2e5a87c3022d38dd5bbfea43540cbdaa", "e01b28388cfd2bbde87baf00c6946e2d", "d653b1bc8a54c633f0922893badc5bee",
    "95888400f19a133b10d8c4323f7e2065", "55196a6e045b3dbf8a708273ebf56864"
}

remove_tags = {
    "احراز- احراز خودکار", "احراز- ثبت نام تکراری", "احراز- هویت", "احراز-آدرس",
    "احراز-اطلاعات بانکی", "احراز-تصویر احراز", "احراز-تغییرهمراه", "احراز - سطح 3",
    "احراز - شرایط ثبت نام", "احراز -سطح 2"
}

YOUR_EMAIL = "alireza.oshani.1989am@gmail.com"
YOUR_NAME = "Powered by Alireza Oshani"

class WorkerSignals(QObject):
    success = pyqtSignal(str)
    error = pyqtSignal(str)
    progress_start = pyqtSignal()
    progress_stop = pyqtSignal()

class GlassTab(QWidget):
    def __init__(self, tab_title="SLA"):
        super().__init__()
        self.tab_title = tab_title
        self.signals = WorkerSignals()
        self.start_button = None  # Store the start button
        self.init_ui()

    def init_ui(self):
        title = QLabel(self.tab_title)
        title.setStyleSheet("font-size:23px;font-weight:bold;color:#fff;")
        title.setAlignment(Qt.AlignCenter)

        edit_width = 320
        button_width = 320
        edit_policy = QSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        button_policy = QSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.input_edit = QLineEdit()
        self.input_edit.setPlaceholderText("مسیر فایل ورودی (TSV یا Excel)")
        self.input_edit.setStyleSheet(
            "background:rgba(255,255,255,0.08);color:#fff;padding:7px;border-radius:7px;"
        )
        self.input_edit.setAlignment(Qt.AlignCenter)
        self.input_edit.setFixedWidth(edit_width)
        self.input_edit.setSizePolicy(edit_policy)

        btn_browse_input = QPushButton("انتخاب فایل")
        btn_browse_input.clicked.connect(self.browse_input)
        btn_browse_input.setStyleSheet(
            "background:rgba(50,50,60,0.7);color:#fff;border-radius:7px;padding:7px;"
        )
        btn_browse_input.setFixedWidth(button_width)
        btn_browse_input.setSizePolicy(button_policy)

        self.output_edit = QLineEdit()
        self.output_edit.setPlaceholderText("مسیر ذخیره فایل خروجی (Excel)")
        self.output_edit.setStyleSheet(
            "background:rgba(255,255,255,0.08);color:#fff;padding:7px;border-radius:7px;"
        )
        self.output_edit.setAlignment(Qt.AlignCenter)
        self.output_edit.setFixedWidth(edit_width)
        self.output_edit.setSizePolicy(edit_policy)

        btn_browse_output = QPushButton("انتخاب مسیر")
        btn_browse_output.clicked.connect(self.browse_output)
        btn_browse_output.setStyleSheet(
            "background:rgba(50,50,60,0.7);color:#fff;border-radius:7px;padding:7px;"
        )
        btn_browse_output.setFixedWidth(button_width)
        btn_browse_output.setSizePolicy(button_policy)

        self.encoding_combo = QComboBox()
        self.encoding_combo.addItems(['auto-detect', 'utf-8', 'utf-16', 'latin-1', 'cp1252'])
        self.encoding_combo.setStyleSheet(
            "background:rgba(255,255,255,0.12);color:#fff;border-radius:7px;"
        )

        self.start_button = QPushButton("شروع پردازش")
        self.start_button.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #39d353, stop:1 #159957);"
            "color:#fff;font-weight:bold;border-radius:10px;font-size:18px;padding:13px 0px;"
        )
        self.start_button.setFixedWidth(260)
        self.start_button.setMinimumHeight(38)
        self.start_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.start_button.clicked.connect(self.start_processing)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setStyleSheet("QProgressBar {border-radius:7px;}")
        self.progress.setMaximum(0)  # Indeterminate mode

        self.status_label = QLabel("آماده به کار")
        self.status_label.setStyleSheet("color:gray;font-size:13px;")
        self.status_label.setAlignment(Qt.AlignCenter)

        vbox = QVBoxLayout()
        vbox.addWidget(title)
        vbox.addSpacing(22)

        # Input section
        input_vbox = QVBoxLayout()
        input_vbox.addWidget(self.input_edit)
        input_vbox.addSpacing(10)
        input_vbox.addWidget(btn_browse_input)
        input_vbox.setAlignment(Qt.AlignCenter)

        # Output section
        output_vbox = QVBoxLayout()
        output_vbox.addWidget(self.output_edit)
        output_vbox.addSpacing(10)
        output_vbox.addWidget(btn_browse_output)
        output_vbox.setAlignment(Qt.AlignCenter)

        # Add input and output sections to main layout
        vbox.addLayout(input_vbox)
        vbox.addSpacing(16)
        vbox.addLayout(output_vbox)
        vbox.addSpacing(22)

        # Encoding section
        h_enc = QHBoxLayout()
        lbl_encoding = QLabel("انکودینگ:")
        lbl_encoding.setStyleSheet("color:#fff;font-size:13px;")
        h_enc.addStretch()
        h_enc.addWidget(lbl_encoding)
        h_enc.addSpacing(5)
        h_enc.addWidget(self.encoding_combo)
        h_enc.addStretch()
        vbox.addLayout(h_enc)

        vbox.addSpacing(36)
        vbox.addWidget(self.start_button, alignment=Qt.AlignCenter)
        vbox.addSpacing(10)
        vbox.addWidget(self.progress)
        vbox.addWidget(self.status_label)
        vbox.addStretch()

        container = QWidget()
        container.setLayout(vbox)
        container.setStyleSheet("background:transparent;")
        lay = QVBoxLayout(self)
        lay.addWidget(container)

        # Connect signals
        self.signals.success.connect(self.show_success)
        self.signals.error.connect(self.show_error)
        self.signals.progress_start.connect(self.start_progress)
        self.signals.progress_stop.connect(self.stop_progress)

    def browse_input(self):
        fileName, _ = QFileDialog.getOpenFileName(self, "انتخاب فایل ورودی", "", "TSV/Excel Files (*.tsv *.xlsx *.xls)")
        if fileName:
            self.input_edit.setText(fileName)

    def browse_output(self):
        fileName, _ = QFileDialog.getSaveFileName(self, "مسیر ذخیره اکسل", "", "Excel Files (*.xlsx)")
        if fileName:
            self.output_edit.setText(fileName)

    def start_progress(self):
        self.progress.setVisible(True)
        self.status_label.setText("در حال پردازش...")
        self.status_label.setStyleSheet("color:blue;font-size:13px;")

    def stop_progress(self):
        self.progress.setVisible(False)
        self.start_button.setEnabled(True)

    def show_success(self, message):
        self.status_label.setText(message)
        self.status_label.setStyleSheet("color:green;font-size:13px;")
        QApplication.instance().beep()

    def show_error(self, message):
        self.status_label.setText(message)
        self.status_label.setStyleSheet("color:red;font-size:13px;")
        QApplication.instance().beep()

    def start_processing(self):
        input_path = self.input_edit.text()
        output_path = self.output_edit.text()
        if not (input_path.lower().endswith('.tsv') or input_path.lower().endswith(('.xlsx', '.xls'))):
            self.show_error("لطفاً یک فایل با فرمت TSV یا اکسل انتخاب کنید")
            return
        if not output_path:
            self.show_error("لطفاً مسیر ذخیره فایل را مشخص کنید")
            return
        self.start_button.setEnabled(False)
        self.signals.progress_start.emit()
        threading.Thread(target=self.process_files, args=(input_path, output_path), daemon=True).start()

    def process_files(self, input_path, output_path):
        try:
            chardet, pd, load_workbook, PatternFill, Alignment = import_processing_libs()

            if input_path.lower().endswith('.tsv'):
                encoding = self.encoding_combo.currentText()
                if encoding == 'auto-detect':
                    with open(input_path, 'rb') as f:
                        rawdata = f.read(10000)
                        result = chardet.detect(rawdata)
                        encoding = result['encoding'] or 'utf-8'
                try:
                    df = pd.read_csv(input_path, sep='\t', encoding=encoding, usecols=self.get_required_columns())
                except UnicodeDecodeError:
                    for alt_enc in ['utf-8', 'latin-1', 'cp1252']:
                        try:
                            df = pd.read_csv(input_path, sep='\t', encoding=alt_enc, usecols=self.get_required_columns())
                            break
                        except:
                            continue
                    else:
                        raise ValueError("خطا در تشخیص Encoding")
            else:
                df = pd.read_excel(input_path, usecols=self.get_required_columns())

            if self.tab_title == "SLA":
                self.process_sla(df, output_path, pd, load_workbook, PatternFill, Alignment)
            elif self.tab_title == "FCR":
                self.process_fcr(df, output_path, pd)
            elif self.tab_title == "Abandonment":
                self.process_abandonment(df, output_path, pd, load_workbook, PatternFill, Alignment)

            self.signals.success.emit(f"پردازش {self.tab_title} با موفقیت انجام شد!")
        except Exception as e:
            self.signals.error.emit(f"خطا در پردازش {self.tab_title}: {str(e)}")
        finally:
            self.signals.progress_stop.emit()

    def get_required_columns(self):
        if self.tab_title == "SLA":
            return ['chat creation date Asia/Tehran', 'queue duration in seconds', 'last operator id', 'pre chat: موضوع سوال']
        elif self.tab_title == "FCR":
            return ["conferenceId", "visitor email", "last operator id", "pre chat: موضوع سوال", "tag 1"]
        else:  # Abandonment
            return ["conference ID", "queue start date Asia/Tehran", "pre chat: موضوع سوال"]

    def process_sla(self, df, output_path, pd, load_workbook, PatternFill, Alignment):
        df['chat creation date Asia/Tehran'] = pd.to_datetime(df['chat creation date Asia/Tehran'])
        df['hour'] = df['chat creation date Asia/Tehran'].dt.hour
        df['queue duration in seconds'] = df['queue duration in seconds'].fillna(0)
        df['SLA'] = df['queue duration in seconds'].apply(lambda x: 'کمتر' if x <= 20 else 'بیشتر از 20 ثانیه')
        counts_less = df.groupby('hour')['SLA'].apply(lambda x: (x == 'کمتر').sum()).reindex(range(24), fill_value=0)
        counts_total = df.groupby('hour')['SLA'].count().reindex(range(24), fill_value=0)
        morning_mask = df['chat creation date Asia/Tehran'].dt.hour.between(8, 15)
        evening_mask = df['chat creation date Asia/Tehran'].dt.hour.between(16, 23)
        night_mask = df['chat creation date Asia/Tehran'].dt.hour.between(0, 7)
        awt_morning = df.loc[morning_mask, 'queue duration in seconds'].mean()
        awt_evening = df.loc[evening_mask, 'queue duration in seconds'].mean()
        awt_night = df.loc[night_mask, 'queue duration in seconds'].mean()
        awt_total = df['queue duration in seconds'].mean()
        count_morning = df.loc[morning_mask].shape[0]
        count_evening = df.loc[evening_mask].shape[0]
        count_night = df.loc[night_mask].shape[0]
        count_total = df.shape[0]
        awt_morning = awt_morning if pd.notna(awt_morning) else 0
        awt_evening = awt_evening if pd.notna(awt_evening) else 0
        awt_night = awt_night if pd.notna(awt_night) else 0
        awt_total = awt_total if pd.notna(awt_total) else 0
        df = df.drop(columns=['hour'])
        df.to_excel(output_path, index=False)
        wb = load_workbook(output_path)
        ws = wb.active
        date_col = list(df.columns).index('chat creation date Asia/Tehran') + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=date_col, max_col=date_col):
            for cell in row:
                cell.number_format = 'h'
        ws_sla = wb.create_sheet("Sla Table")
        night_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        morning_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        evening_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        ws_sla.cell(row=1, column=1, value="shifts").alignment = align_center
        ws_sla.merge_cells(start_row=1, start_column=2, end_row=1, end_column=9)
        cell_night = ws_sla.cell(row=1, column=2, value="Night")
        cell_night.fill = night_fill
        cell_night.alignment = align_center
        ws_sla.merge_cells(start_row=1, start_column=10, end_row=1, end_column=17)
        cell_morning = ws_sla.cell(row=1, column=10, value="Morning")
        cell_morning.fill = morning_fill
        cell_morning.alignment = align_center
        ws_sla.merge_cells(start_row=1, start_column=18, end_row=1, end_column=25)
        cell_evening = ws_sla.cell(row=1, column=18, value="Evening")
        cell_evening.fill = evening_fill
        cell_evening.alignment = align_center
        ws_sla.cell(row=2, column=1, value="time").alignment = align_center
        for hr in range(24):
            col_index = hr + 2
            ws_sla.cell(row=2, column=col_index, value=f"{hr:02d}:00").alignment = align_center
        ws_sla.cell(row=3, column=1, value="چت های کمتر از 20 ثانیه").alignment = align_center
        for hr in range(24):
            col_index = hr + 2
            ws_sla.cell(row=3, column=col_index, value=counts_less.get(hr, 0)).alignment = align_center
        ws_sla.cell(row=4, column=1, value="تعداد کل ورودی چت").alignment = align_center
        for hr in range(24):
            col_index = hr + 2
            ws_sla.cell(row=4, column=col_index, value=counts_total.get(hr, 0)).alignment = align_center
        ws_awt = wb.create_sheet("awt")
        ws_awt.cell(row=1, column=1, value="shifts").alignment = align_center
        ws_awt.cell(row=1, column=2, value="Morning").alignment = align_center
        ws_awt.cell(row=1, column=3, value="Evening").alignment = align_center
        ws_awt.cell(row=1, column=4, value="Night").alignment = align_center
        ws_awt.cell(row=1, column=5, value="Total").alignment = align_center
        ws_awt.cell(row=2, column=1, value="Awt").alignment = align_center
        ws_awt.cell(row=2, column=2, value=round(awt_morning, 2)).alignment = align_center
        ws_awt.cell(row=2, column=3, value=round(awt_evening, 2)).alignment = align_center
        ws_awt.cell(row=2, column=4, value=round(awt_night, 2)).alignment = align_center
        ws_awt.cell(row=2, column=5, value=round(awt_total, 2)).alignment = align_center
        ws_awt.cell(row=3, column=1, value="تعداد چت").alignment = align_center
        ws_awt.cell(row=3, column=2, value=count_morning).alignment = align_center
        ws_awt.cell(row=3, column=3, value=count_evening).alignment = align_center
        ws_awt.cell(row=3, column=4, value=count_night).alignment = align_center
        ws_awt.cell(row=3, column=5, value=count_total).alignment = align_center
        wb.save(output_path)

    def process_fcr(self, df, output_path, pd):
        df = df[~df["last operator id"].isin(remove_ids)]
        total_count = df["conferenceId"].count()
        summary = pd.DataFrame({"عنوان": ["تعداد کل"], "مقدار": [total_count]})
        df = df[df["tag 1"] != "لفت"]
        df['fcr_value'] = df.groupby(['visitor email', 'tag 1'])['conferenceId'].transform('count')
        multi_visit_count = (df['fcr_value'] > 1).sum()
        df.drop('fcr_value', axis=1, inplace=True)
        formulas = [f"=COUNTIFS(B:B,B{i+2},E:E,E{i+2})" for i in range(len(df))]
        df["fcr"] = formulas
        summary = pd.concat([summary, pd.DataFrame({"عنوان": ["چند بار مراجعه"], "مقدار": [multi_visit_count]})], ignore_index=True)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="داده‌ها")
            summary.to_excel(writer, index=False, sheet_name="خلاصه")

    def process_abandonment(self, df, output_path, pd, load_workbook, PatternFill, Alignment):
        df["queue start date Asia/Tehran"] = pd.to_datetime(df["queue start date Asia/Tehran"])
        df["queue start date Asia/Tehran"] = df["queue start date Asia/Tehran"].dt.hour.apply(lambda h: f"{h:02d}:00")
        df['hour'] = pd.to_datetime(df["queue start date Asia/Tehran"], format="%H:%M").dt.hour
        counts = df.groupby('hour').size().reindex(range(24), fill_value=0)
        df = df.drop(columns=['hour'])
        df.to_excel(output_path, index=False)
        wb = load_workbook(output_path)
        ws_qab = wb.create_sheet("Queue Abandonment Table")
        align_center = Alignment(horizontal="center", vertical="center")
        ws_qab.cell(row=1, column=1, value="time").alignment = align_center
        for hr in range(24):
            ws_qab.cell(row=1, column=hr+2, value=f"{hr:02d}:00").alignment = align_center
        ws_qab.cell(row=2, column=1, value="تعداد").alignment = align_center
        for hr in range(24):
            ws_qab.cell(row=2, column=hr+2, value=counts.get(hr, 0)).alignment = align_center
        wb.save(output_path)

class GlassMainWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.radius = 32
        self.setStyleSheet("background:transparent;")
        self.setObjectName("glassMain")

        self.titlebar = QFrame(self)
        self.titlebar.setFixedHeight(58)
        self.titlebar.setStyleSheet("QFrame {background:transparent;}")
        hbox = QHBoxLayout(self.titlebar)
        hbox.setContentsMargins(18, 0, 10, 0)
        hbox.addStretch(1)
        self.title = QLabel("SLA & FCR Processor")
        self.title.setStyleSheet("font-size:20px;color:#f4f4f4;font-weight:bold;")
        self.title.setAlignment(Qt.AlignCenter)
        hbox.addWidget(self.title)
        hbox.addStretch(1)
        self.min_btn = QPushButton("➖")
        self.min_btn.setFixedSize(28, 28)
        self.min_btn.setStyleSheet(
            "QPushButton{border:none;color:#fff;background:rgba(120,120,130,0.20);border-radius:14px;}"
            "QPushButton:hover{background:#23aaff;color:#fff;}"
        )
        self.min_btn.setCursor(QCursor(Qt.PointingHandCursor))
        hbox.addWidget(self.min_btn)
        self.close_btn = QPushButton("❌")
        self.close_btn.setFixedSize(28, 28)
        self.close_btn.setStyleSheet(
            "QPushButton{border:none;color:#fff;background:rgba(120,60,60,0.25);border-radius:14px;}"
            "QPushButton:hover{background:#ff324c;color:#fff;}"
        )
        self.close_btn.setCursor(QCursor(Qt.PointingHandCursor))
        hbox.addWidget(self.close_btn)
        self.min_btn.clicked.connect(lambda: self.window().showMinimized())
        self.close_btn.clicked.connect(lambda: self.window().close())

        tabs_container = QWidget()
        tabs_hbox = QHBoxLayout(tabs_container)
        tabs_hbox.setContentsMargins(26, 0, 26, 0)
        tabs_hbox.setSpacing(0)
        tabs_hbox.addStretch(1)

        self.tab_widget = QTabWidget(self)
        self.tab_widget.setUsesScrollButtons(False)  # Disable scroll buttons
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane { border: 0; background:transparent; }
            QTabBar::tab { 
                background: rgba(28,28,34,200);
                color: #fff;
                min-width:110px; min-height:26px; 
                border-radius: 13px; font-size:16px; 
                margin: 7px 8px 0px 8px; padding:4px 15px;
            }
            QTabBar::tab:selected { background: #11121b; color: #39d353; }
            QTabBar::scroller { width: 0; }  /* Hide scroll buttons */
            QTabBar QToolButton { width: 0; height: 0; }  /* Hide scroll arrows */
        """)
        self.sla_tab = GlassTab("SLA")
        self.fcr_tab = GlassTab("FCR")
        self.ab_tab = GlassTab("Abandonment")
        self.tab_widget.addTab(self.sla_tab, "SLA")
        self.tab_widget.addTab(self.fcr_tab, "FCR")
        self.tab_widget.addTab(self.ab_tab, "Abandonment")

        tabs_hbox.addWidget(self.tab_widget)
        tabs_hbox.addStretch(1)

        self.statusbar = QFrame(self)
        self.statusbar.setFixedHeight(36)
        self.statusbar.setStyleSheet("background:transparent;")
        status_layout = QHBoxLayout(self.statusbar)
        status_layout.setContentsMargins(14, 0, 14, 0)
        name_label = QLabel(YOUR_NAME)
        name_label.setStyleSheet("color:#b3ffa6;font-size:12px;font-weight:bold;")
        email_label = QLabel(YOUR_EMAIL)
        email_label.setStyleSheet("color:#eee;font-size:12px;")
        status_layout.addWidget(name_label)
        status_layout.addStretch(1)
        status_layout.addWidget(email_label)

        vbox = QVBoxLayout(self)
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(0)
        vbox.addWidget(self.titlebar)
        vbox.addWidget(tabs_container)
        vbox.addWidget(self.statusbar)

    def paintEvent(self, event):
        painter = QPainter(self)
        color = QColor(10, 10, 18, 225)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setBrush(QBrush(color))
        painter.setPen(Qt.NoPen)
        rect = self.rect()
        painter.drawRoundedRect(rect, self.radius, self.radius)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_pos = event.globalPos() - self.parent().frameGeometry().topLeft()
            event.accept()

    def mouseMoveEvent(self, event):
        if hasattr(self, '_drag_pos') and event.buttons() == Qt.LeftButton:
            self.parent().move(event.globalPos() - self._drag_pos)
            event.accept()

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SLA & FCR Processor")
        self.setGeometry(350, 120, 580, 630)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Window)
        self.setAttribute(Qt.WA_TranslucentBackground)

        self.glass_card = GlassMainWidget(self)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 16, 24, 16)
        layout.addWidget(self.glass_card)

    def mousePressEvent(self, event):
        self.glass_card.mousePressEvent(event)

    def mouseMoveEvent(self, event):
        self.glass_card.mouseMoveEvent(event)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())